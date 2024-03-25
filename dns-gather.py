from azure.identity import DefaultAzureCredential
from azure.mgmt.compute import ComputeManagementClient
from azure.mgmt.web import WebSiteManagementClient
from azure.mgmt.resource import ResourceManagementClient
from azure.mgmt.network import NetworkManagementClient
from openpyxl import Workbook
from azure.core.exceptions import ResourceNotFoundError
from kubernetes import client, config
from azure.mgmt.containerservice import ContainerServiceClient

from common.aks.network import get_all_service_ips_aks
from common.vm.dns import get_dns_info
from common.network.network import get_ip_addresses

# Initialize Azure credentials
credentials = DefaultAzureCredential()
subscription_id = 'a6cc1a53-c242-42f9-aa16-15a377d21069'

# Initialize AKS client
aks_client = ContainerServiceClient(credentials, subscription_id)
# Retrieve list of AKS clusters
azure_aks_clusters = aks_client.managed_clusters.list()

# Initialize Azure clients
compute_client = ComputeManagementClient(credentials, subscription_id)
network_client = NetworkManagementClient(credentials, subscription_id)
web_client = WebSiteManagementClient(credentials, subscription_id)
resource_client = ResourceManagementClient(DefaultAzureCredential(), subscription_id)

# Create a new Excel workbook
workbook = Workbook()

# Virtual Machines
vm_sheet = workbook.create_sheet("Virtual Machines")
vm_sheet.append(["Resource Group", "Host", "Private IP", "Public IP", "Internal Domain Name Suffix", "Private DNS", "DNS Servers"])
for rg in resource_client.resource_groups.list():
    for vm in compute_client.virtual_machines.list(rg.name):
        for nic_reference in vm.network_profile.network_interfaces:
            try:
                nic = network_client.network_interfaces.get(rg.name, nic_reference.id.split('/')[-1])
                dns_info = get_dns_info(vm, nic.id, network_client)
                if dns_info is not None:
                    custom_dns, private_dns, dns_servers = dns_info
                else:
                    # Assign default values if get_dns_info() returns None
                    custom_dns, private_dns, dns_servers = "N/A", "N/A", "N/A"
                
                private_ip, public_ip = get_ip_addresses(nic, network_client)
                
                # Convert DNS servers to a string
                dns_servers_str = ", ".join(dns_servers) if dns_servers != "N/A" else "N/A"
                
                vm_sheet.append([rg.name, vm.name, private_ip, public_ip, custom_dns, private_dns, dns_servers_str])

            except ResourceNotFoundError as e:
                print(f"Resource not found while retrieving NIC: {e}")
                # Continue to the next iteration
                continue

            except Exception as e:
                print(f"Error occurred while retrieving NIC: {e}")
                # Log the error or handle it as needed
                continue


# AKS Services Information
aks_sheet = workbook.create_sheet("AKS")
aks_sheet.append(["Resource Group", "AKS Server", "Namespace", "Service", "Service Type", "Service IP", "External IP", "Ingress Name", "Ingress Namespace", "Ingress Class", "Ingress Hosts", "Ingress Address"])

# Iterate through each AKS cluster
for cluster in azure_aks_clusters:
    cluster_resource_group = cluster.node_resource_group
    cluster_name = cluster.name
    
    # Retrieve service information for this AKS cluster
    cluster_service_info = get_all_service_ips_aks(cluster_resource_group, [cluster_name])

    # Iterate through the service information for this AKS cluster
    for cluster_name, cluster_info in cluster_service_info.items():
        for aks_namespace, services in cluster_info.items():
            for service_name, service_details in services.items():
                aks_sheet.append([
                    cluster_resource_group,
                    cluster_name,  # Assuming cluster_name is a string
                    aks_namespace,
                    service_name,
                    service_details["Type"],
                    service_details["IP"],
                    service_details["External IP"],  # New field
                    service_details.get("Ingress Name", ""),  # New field
                    service_details.get("Ingress Namespace", ""),  # New field
                    service_details.get("Ingress Class", ""),  # New field
                    ",".join(service_details.get("Ingress Hosts", [])),  # New field, if hosts are a list
                    ",".join(service_details.get("Ingress Address", []))  # New field, if addresses are a list
                ])


# App Services
app_services_sheet = workbook.create_sheet("App Services")
app_services_sheet.append(["Resource Group", "App Services", "Default Domain", "Custom Domains"])
for rg in resource_client.resource_groups.list():
    for site in web_client.web_apps.list_by_resource_group(rg.name):
        app_services_sheet.append([rg.name, site.name, site.default_host_name, ", ".join(site.enabled_host_names)])

# Save the workbook
workbook.save("azure_resources.xlsx")
